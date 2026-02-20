import streamlit as st
import pandas as pd
import os
import re
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO

# --- 페이지 설정 및 프리미엄 스타일 ---
st.set_page_config(page_title="Keyword Master Pro", page_icon="📈", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
    .stApp { background-color: #F8FAFC; }
    .main-header { font-size: 2.2rem; font-weight: 700; color: #1E293B; margin-bottom: 0.5rem; text-align: center; }
    .sub-header { font-size: 1rem; color: #64748B; margin-bottom: 2rem; text-align: center; }
    .metric-card { background: white; padding: 20px; border-radius: 12px; border: 1px solid #E2E8F0; text-align: center; box-shadow: 0 1px 3px rgba(0,0,0,0.1); }
    .metric-val { font-size: 1.8rem; font-weight: 700; color: #2563EB; }
    .metric-label { font-size: 0.8rem; color: #64748B; font-weight: 600; margin-top: 5px; }
    .stButton>button { width: 100%; background-color: #2563EB; color: white; border-radius: 8px; padding: 0.7rem; font-weight: 600; border: none; }
</style>
""", unsafe_allow_html=True)

# --- 엑셀 스타일링 함수 (Colab 버전과 100% 일치) ---
def get_styled_excel(writer, results, highlight_map, month_names):
    h_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    b_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    headers = ["키워드"] + [f"{int(m[2:])}월" for m in month_names]
    
    for sn in ['사계절키워드', '시즌키워드', '성장키워드']:
        data = results[sn]
        extra = ["평균검색량", "규모등급", "안정성"] if sn == "사계절키워드" else (["피크월", "규모등급", "소싱타이밍"] if sn == "시즌키워드" else ["상승구간", "규모등급", "성장유형"])
        df = pd.DataFrame(data if data else [["해당 키워드 없음"] + ["-"]*(len(headers)+len(extra)-1)], columns=headers + extra)
        df.to_excel(writer, sheet_name=sn, index=False)
        
        ws = writer.sheets[sn]
        for col in ws.columns:
            max_len = 0
            for cell in col:
                if cell.value:
                    v_str = str(cell.value)
                    cur_len = sum(2 if ord(c) > 128 else 1 for c in v_str)
                    if cur_len > max_len: max_len = cur_len
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
            for cell in col:
                cell.alignment = Alignment(vertical='center', horizontal='center')
                if cell.row == 1: cell.fill, cell.font = h_fill, Font(bold=True)
                if cell.column == 1: cell.alignment = Alignment(vertical='center', horizontal='left')
        
        if sn in highlight_map:
            for r_idx, cols in highlight_map[sn].items():
                for c_idx in cols: ws.cell(row=r_idx + 2, column=c_idx + 1).fill = b_fill

# --- 메인 화면 레이아웃 ---
st.markdown('<p class="main-header">📈 Keyword Master Pro</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-header">아이템스카우트 데이터 정밀 분석 프리미엄 대시보드</p>', unsafe_allow_html=True)

# 대시보드 분석 기준 안내
with st.expander("💡 분석 기준 안내 (자세히 보려면 클릭)"):
    st.markdown("""
    <div style="background-color: white; padding: 10px; border-radius: 10px;">
    <table style="width:100%; border-collapse: collapse; font-size: 14px; text-align: center;">
        <tr style="background-color: #F1F5F9; font-weight: bold; border-bottom: 2px solid #E2E8F0;">
            <th style="padding: 12px;">구분</th>
            <th style="padding: 12px;">지표명</th>
            <th style="padding: 12px;">기준점 (수치)</th>
            <th style="padding: 12px;">의미</th>
        </tr>
        <tr style="border-bottom: 1px solid #F1F5F9;">
            <td style="padding: 10px; font-weight: 600;">공통</td>
            <td style="padding: 10px;">규모 등급</td>
            <td style="padding: 10px;">Gold: 1만↑ / Silver: 5천↑ / Bronze: 3천↑</td>
            <td style="padding: 10px; color: #64748B;">시장의 크기를 한눈에 파악</td>
        </tr>
        <tr style="border-bottom: 1px solid #F1F5F9;">
            <td style="padding: 10px; font-weight: 600;">사계절</td>
            <td style="padding: 10px;">안정성</td>
            <td style="padding: 10px;">A+: 변동 10%↓ / A: 20%↓ / B: 30%↓</td>
            <td style="padding: 10px; color: #64748B;">수요가 얼마나 기복 없이 탄탄한가</td>
        </tr>
        <tr style="border-bottom: 1px solid #F1F5F9;">
            <td style="padding: 10px; font-weight: 600;">시즌</td>
            <td style="padding: 10px;">소싱타이밍</td>
            <td style="padding: 10px;">피크월 대비 4개월 전</td>
            <td style="padding: 10px; color: #64748B;">노출 선점을 위한 준비 시점</td>
        </tr>
        <tr>
            <td style="padding: 10px; font-weight: 600;">성장</td>
            <td style="padding: 10px;">성장 유형</td>
            <td style="padding: 10px;">폭발: 월평균 20%↑ / 꾸준: 월평균 5~20%↑</td>
            <td style="padding: 10px; color: #64748B;">상승 곡선의 가파른 정도 확인</td>
        </tr>
    </table>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("<br>", unsafe_allow_html=True)

col1, col2 = st.columns([1, 2], gap="large")

with col1:
    st.markdown('<div style="background: white; padding: 20px; border-radius: 12px; border: 1px solid #E2E8F0;">', unsafe_allow_html=True)
    st.subheader("⚙️ 분석 설정")
    target_cat_input = st.text_input("분석 카테고리명", value="실버용품")
    start_yymm = st.text_input("시작 월 (YYMM)", value="2601")
    st.markdown("---")
    st.write("📂 **데이터 업로드** (엑셀 파일 선택)")
    
    # 파일 업로더 초기화용 키 설정
    if 'file_key' not in st.session_state:
        st.session_state['file_key'] = 0

    uploaded = st.file_uploader("", accept_multiple_files=True, label_visibility="collapsed", key=f"uploader_{st.session_state['file_key']}")
    
    # 버튼 나란히 배치
    btn_col1, btn_col2 = st.columns(2)
    with btn_col1:
        analyze_btn = st.button("🚀 정밀 분석 시작")
    with btn_col2:
        if st.button("🗑️ 일괄 삭제"):
            st.session_state['file_key'] += 1
            st.rerun()
    st.markdown('</div>', unsafe_allow_html=True)

with col2:
    if analyze_btn:
        if not uploaded:
            st.error("파일을 먼저 업로드해주세요!")
        else:
            with st.spinner("전문 엔진 분석 중..."):
                try:
                    # 1. 파일 처리
                    files = sorted(uploaded, key=lambda x: x.name)
                    year, m_start = int(start_yymm[:2]), int(start_yymm[2:])
                    month_names = []
                    for i in range(len(files)):
                        cm, cy = m_start + i, year
                        while cm > 12: cm -= 12; cy += 1
                        month_names.append(f"{cy:02d}{cm:02d}")
                    
                    kw_map = {}
                    first_df = pd.read_excel(files[0])
                    
                    # 지능형 뎁스 판별 (특수문자 무시 매칭)
                    def normalize(text):
                        return re.sub(r'[^a-zA-Z0-9가-힣]', '', str(text)) if text else ""
                    
                    norm_target = normalize(target_cat_input)
                    t_depth, act_cat = None, target_cat_input
                    
                    # 모든 행을 뒤져서 특수문자 무시하고 매칭되는 뎁스 찾기
                    sample_cats = first_df['대표 카테고리'].dropna().unique()[:100]
                    for cat_str in sample_cats:
                        parts = [p.strip() for p in str(cat_str).split('>')]
                        for idx, p in enumerate(parts):
                            if normalize(p) == norm_target:
                                t_depth = idx + 1
                                act_cat = p # 실제 데이터에 적힌 정확한 명칭 보관
                                break
                        if t_depth: break
                    
                    if not t_depth:
                        st.error(f"❌ 카테고리 [{target_cat_input}]를 데이터에서 찾을 수 없습니다.")
                        st.info("데이터에 포함된 카테고리명을 정확히 입력하거나, '스포츠/레저'처럼 슬래시를 포함해 보세요.")
                        st.stop()
                    
                    st.info(f"🔍 분석 기준: {t_depth}차 카테고리 [{act_cat}]")

                    # 2. 데이터 수집
                    for idx, f in enumerate(files):
                        df = pd.read_excel(f)
                        # 검색수 숫자 변환 (비수치 데이터 '-' 등 처리)
                        df['총 검색수'] = pd.to_numeric(df['총 검색수'], errors='coerce').fillna(0)
                        
                        # 뎁스에 맞는 카테고리 추출 및 노멀라이즈 비교
                        def get_match(cat_full, target_depth, target_norm):
                            ps = [p.strip() for p in str(cat_full).split('>')]
                            if len(ps) >= target_depth and normalize(ps[target_depth-1]) == target_norm:
                                return True
                            return False

                        for _, row in df.iterrows():
                            if get_match(row['대표 카테고리'], t_depth, norm_target):
                                kw = str(row['키워드']).strip()
                                if kw not in kw_map: kw_map[kw] = [0] * len(files)
                                kw_map[kw][idx] = float(row['총 검색수'])
                    
                    # 3. 분류 (Colab 엔진과 100% 동일)
                    results = {'사계절키워드': [], '시즌키워드': [], '성장키워드': []}
                    highlight_map, total_found = {}, 0
                    for kw, counts in kw_map.items():
                        valid = [c for c in counts if c > 0]
                        if len(valid) < 2: continue
                        avg = sum(valid) / len(valid)
                        if avg < 3000: continue
                        total_found += 1
                        grade = "Gold" if avg >= 10000 else ("Silver" if avg >= 5000 else "Bronze")
                        
                        vars = [abs(counts[i+1]-counts[i])/counts[i] for i in range(len(counts)-1) if counts[i]>0 and counts[i+1]>0]
                        if vars and all(v < 0.3 for v in vars) and all(c > 0 for c in counts):
                            m_v = max(vars); stab = "A+" if m_v < 0.1 else ("A" if m_v < 0.2 else "B")
                            results['사계절키워드'].append([kw] + counts + [round(avg), grade, stab])
                            continue
                        if max(counts) >= avg * 1.3:
                            p_idx = counts.index(max(counts))
                            results['시즌키워드'].append([kw] + counts + [f"{int(month_names[p_idx][2:])}월", grade, f"{int(month_names[(p_idx-4)%len(month_names)][2:])}월"])
                            highlight_map.setdefault('시즌키워드', {})[len(results['시즌키워드'])-1] = [p_idx + 1]
                            continue
                        g_ps, g_cs, s_idx, cnt, t_rate, steps = [], [], -1, 0, 0, 0
                        for i in range(len(counts)-1):
                            if counts[i] > 0 and counts[i+1] >= counts[i] * 1.05:
                                t_rate += (counts[i+1]-counts[i])/counts[i]; steps += 1
                                if s_idx == -1: s_idx = i; cnt = 2
                                else: cnt += 1
                            else:
                                if cnt >= 3:
                                    g_ps.append(f"{int(month_names[s_idx][2:])}→{int(month_names[i][2:])}월")
                                    g_cs.extend(range(s_idx + 1, i + 2))
                                s_idx, cnt = -1, 0
                        if cnt >= 3:
                            g_ps.append(f"{int(month_names[s_idx][2:])}→{int(month_names[-1][2:])}월")
                            g_cs.extend(range(s_idx + 1, len(counts) + 1))
                        if g_ps:
                            results['성장키워드'].append([kw] + counts + [", ".join(g_ps), grade, "폭발" if (t_rate/steps) >= 0.2 else "꾸준"])
                            highlight_map.setdefault('성장키워드', {})[len(results['성장키워드'])-1] = list(set(g_cs))

                    # 4. 결과 출력
                    st.balloons()
                    st.success(f"[{act_cat}] 분석 완료!")
                    d1, d2, d3 = st.columns(3)
                    d1.markdown(f'<div class="metric-card"><div class="metric-val">{len(results["사계절키워드"])}</div><div class="metric-label">사계절 스테디</div></div>', unsafe_allow_html=True)
                    d2.markdown(f'<div class="metric-card"><div class="metric-val">{len(results["시즌키워드"])}</div><div class="metric-label">시즌 트렌드</div></div>', unsafe_allow_html=True)
                    d3.markdown(f'<div class="metric-card"><div class="metric-val">{len(results["성장키워드"])}</div><div class="metric-label">급성장 유망</div></div>', unsafe_allow_html=True)
                    
                    output = BytesIO()
                    with pd.ExcelWriter(output, engine='openpyxl') as writer:
                        get_styled_excel(writer, results, highlight_map, month_names)
                    
                    st.markdown("---")
                    file_name = f"키워드분석_{act_cat}_{month_names[0]}~{month_names[-1]}.xlsx"
                    st.download_button(
                        label="📥 정밀 분석 리포트 다운로드 (Excel)",
                        data=output.getvalue(),
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"오류: {e}")
    else:
        st.info("왼쪽 대시보드에서 파일을 업로드하고 버튼을 눌러주세요.")